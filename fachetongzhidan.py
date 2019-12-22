import openpyxl,arrow
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.utils import get_column_letter, column_index_from_string

fache_data = {
<<<<<<< HEAD
    '计划发车日期': '2019.12.04',
=======
    '计划发车日期': '2019.11.20',
>>>>>>> b2587a0ece98316c4527002f4faff68a5ebaffdb
    '发货单位': '盛宝公司',  #可复制以下公司名称快捷替换
    # 忠诚 中农  阿徐 创业 稼贾福 联益 瑞和泰 庄海钦 穗易通  鹏润  远泰  王台  张安礼   张安芬
    '货物': '大米',
    '到站': '白市驿', #可复制以下到站快捷替换
    # 龙潭寺  白市驿  安顺  西乡  治江  拉萨

    # 编织袋   纸箱
    '规格1': 20,
    '包装1': '编织袋',
<<<<<<< HEAD
    '数量1': 250,
=======
    '数量1': 2392,
>>>>>>> b2587a0ece98316c4527002f4faff68a5ebaffdb

    '规格2': 10,
    '包装2': '编织袋',
    '数量2': 4700,

    '规格3': 20,
    '包装3': '纸箱',
    '数量3': 350,

    '规格4': '',
    '包装4': '',
    '数量4': 0,

    '规格5': '',
    '包装5': '',
    '数量5': 0,
}

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

now = arrow.now().format('YYYY年MM月DD日')
date = now

sheet.merge_cells('A1:H1')
sheet.merge_cells('G1:H1')
sheet['A1'].value = '装车通知单————by业务组('+ date + ')'
sheet['A2'].value = '计划装车日期'
sheet['B2'].value = '到站'
sheet['C2'].value = '货物'
sheet['D2'].value = '规格(KG)'
sheet['E2'].value = '包装'
sheet['F2'].value = '件数'
sheet['G2'].value = '重量(KG)'
sheet['H2'].value = '发货单位'

sheet.merge_cells('A3:A7')
sheet.merge_cells('B3:B7')
sheet.merge_cells('C3:C7')
sheet.merge_cells('H3:H7')

if not fache_data['计划发车日期'] or fache_data['计划发车日期'] == '待定':
    sheet['A3'].value = '待定'
else:
    sheet['A3'].value = arrow.get(fache_data['计划发车日期']).format('YYYY年MM月DD日')

sheet['H3'].value = fache_data['发货单位']
sheet['B3'].value = fache_data['到站']
sheet['C3'].value = fache_data['货物']

#如果规格或数量为空，则替换为0
list1 = [fache_data['规格1'], fache_data['规格2'],fache_data['规格3'],fache_data['规格4'],fache_data['规格5'],
        fache_data['数量1'], fache_data['数量2'], fache_data['数量3'],fache_data['数量4'],fache_data['数量5']]

list2 = [sheet['D3'],sheet['D4'],sheet['D5'],sheet['D6'],sheet['D7'],
         sheet['F3'],sheet['F4'],sheet['F5'],sheet['F6'],sheet['F7']]

for x, y in zip(list1, list2):
    if x:
        y.value = x
    else:
        y.value = 0

sheet['E3'].value = fache_data['包装1']
sheet['E4'].value = fache_data['包装2']
sheet['E5'].value = fache_data['包装3']
sheet['E6'].value = fache_data['包装4']
sheet['E7'].value = fache_data['包装5']
sheet['G3'].value = sheet['D3'].value * sheet['F3'].value
sheet['G4'].value = sheet['D4'].value * sheet['F4'].value
sheet['G5'].value = sheet['D5'].value * sheet['F5'].value
sheet['G6'].value = sheet['D6'].value * sheet['F6'].value
sheet['G7'].value = sheet['D7'].value * sheet['F7'].value

sheet['A8'].value = '合计'
sheet['F8'].value = sum(cell.value for cells in sheet['F3:F5'] for cell in cells)#数量合计
sheet['G8'].value = sum(cell.value for cells in sheet['G3:G5'] for cell in cells)#重量合计

#设置单元格格式
font1 = Font(name='黑体',size=24)
font2 = Font(size=12)
border2 = Border(left=Side('thin'), right=Side('thin'),top=Side('thin'), bottom=Side('thin'))
alignment = Alignment(horizontal='center', vertical='center')

#本代码用于给所有单元格格式化，然后再给第一行标题行单独格式化
def font_cells(s):
    for cells in s.rows:
        for cells in s.columns:
            for cell in cells:
                cell.font = font2
                cell.border = border2
                cell.alignment = alignment

for sheet in wb:
    font_cells(sheet)
for i in range(1, sheet.max_row+1):
    sheet.row_dimensions[i].height = 30
for i in range(1, sheet.max_column+1):
    sheet.column_dimensions[get_column_letter(i)].width = 10
sheet.column_dimensions['A'].width = 16

wb.save('d:\\' + sheet['A3'].value +sheet['H3'].value + sheet['B3'].value + '.xlsx')

